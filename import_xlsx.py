"""
Imports the original "Reaching Unreal.xlsx" workbook into a clean JSON seed
that the React app can load on first run.

The xlsx has two side-by-side mini-tables per sheet (one per user). The
column layout varies week to week. Weights are encoded inside Excel array
formulas like:
    =SUM(Table2[[#This Row],[CS224n Project]:[Research Apps]]*10,
         Table2[[#This Row],[Studying]]*15,
         Table2[[#This Row],[Squash]]*15,
         Table2[[#This Row],[P]]*10,
         Table2[[#This Row],[gym]]*15)

So we parse those formulas to recover per-column weights.
"""
from __future__ import annotations

import json
import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

XLSX_PATH = Path(__file__).parent / "Reaching Unreal.xlsx"
OUT_PATH = Path(__file__).parent / "app" / "src" / "data" / "seed.json"

DAYS = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
USER_LABELS = {"shazly": "El Shazly", "sayed": "El Sayed"}


def parse_weights_from_formula(formula: str | None, headers: list[str]) -> dict[str, float]:
    """Parse the Results array formula to extract weights per column.
    Returns {column_name: weight}. Defaults to 10 for unmentioned columns.
    """
    if not formula:
        return {}
    weights: dict[str, float] = {}
    pattern = re.compile(
        r"\[\[#This Row\],\[([^\]]+)\](?::\[([^\]]+)\])?\]\s*\*\s*(\d+(?:\.\d+)?)"
    )
    for m in pattern.finditer(formula):
        start_col, end_col, w = m.group(1), m.group(2), float(m.group(3))
        if end_col is None:
            weights[start_col.strip()] = w
        else:
            try:
                i_start = headers.index(start_col)
                i_end = headers.index(end_col)
            except ValueError:
                continue
            for i in range(i_start, i_end + 1):
                weights[headers[i].strip()] = w
    return weights


def find_user_table(headers: list, start_marker: str = "Day\\Project") -> list[tuple[int, int]]:
    """Find (start_col_idx, end_col_idx) ranges where each user table lives.
    Returns 1-based column ranges. Includes the Day column and Results column.
    """
    ranges = []
    n = len(headers)
    for i, h in enumerate(headers):
        if h is None:
            continue
        s = str(h).strip().lower()
        if s in {"day\\project", "column1"}:
            # Find Results column to the right within ~20 cols
            for j in range(i + 1, min(n, i + 25)):
                if headers[j] and str(headers[j]).strip().lower() == "results":
                    ranges.append((i + 1, j + 1))  # 1-based inclusive
                    break
    return ranges


def first_saturday_on_or_after(d: datetime) -> datetime:
    # Python: Mon=0 ... Sun=6; Saturday=5
    delta = (5 - d.weekday()) % 7
    return d + timedelta(days=delta)


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    wb_vals = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    weeks: list[dict] = []

    # Determine the first week's start date by reading the date cell from Week 1
    try:
        first_sheet = wb_vals["Week 1"]
        first_date_cell = first_sheet.cell(row=4, column=20).value
        if isinstance(first_date_cell, datetime):
            anchor_start = first_date_cell
        else:
            anchor_start = datetime(2026, 2, 7)
    except KeyError:
        anchor_start = datetime(2026, 2, 7)

    week_sheet_names = [
        s for s in wb.sheetnames if s.lower().startswith("week ")
    ]

    # Sort: "Week 1", "Week 2", ..., "week 13"
    def week_num(name: str) -> int:
        try:
            return int(name.split()[-1])
        except ValueError:
            return 999

    week_sheet_names.sort(key=week_num)

    for sheet_name in week_sheet_names:
        n = week_num(sheet_name)
        sheet_f = wb[sheet_name]
        sheet_v = wb_vals[sheet_name]

        # Header row is row 3 in this workbook
        headers = [c.value for c in sheet_f[3]]
        # Strip trailing None
        while headers and headers[-1] is None:
            headers.pop()

        ranges = find_user_table(headers)

        start_date = anchor_start + timedelta(weeks=(n - 1))
        end_date = start_date + timedelta(days=6)

        tables = []
        # Map known names by ordinal position
        for idx, (col_start, col_end) in enumerate(ranges):
            user_id = list(USER_LABELS.keys())[idx] if idx < len(USER_LABELS) else f"user{idx}"
            user_name = USER_LABELS.get(user_id, f"User {idx + 1}")

            # Column headers: from col_start+1 .. col_end-1 (skip Day and Results)
            col_headers = []
            for c in range(col_start + 1, col_end):
                h = headers[c - 1]
                if h is None:
                    continue
                col_headers.append(str(h).strip())

            # Pull the array formula text from the Results cell at row 4
            results_cell = sheet_f.cell(row=4, column=col_end).value
            formula_text = getattr(results_cell, "text", None) if results_cell else None
            weights = parse_weights_from_formula(formula_text, col_headers)

            # Build columns with sensible defaults if no weight known
            columns = []
            for h in col_headers:
                w = weights.get(h)
                if w is None:
                    # Heuristics for common boolean-style columns
                    lh = h.lower()
                    if any(k in lh for k in ["gym", "squash", "diet", "wake", "sleep", "screentime", "scroll", "shower", "p"]):
                        w = 15
                    else:
                        w = 10
                # Best-effort: treat short/known habits as boolean
                lh = h.lower()
                is_bool = any(
                    k in lh for k in [
                        "gym", "squash", "diet", "wake", "sleep", "screentime",
                        "quit scroll", "shower", "cardio", "p", "rt", "recharging",
                    ]
                ) and len(h) <= 60
                # The 'P' single-letter column is nearly always boolean; CS131 etc. are hours
                col_type = "boolean" if (h.strip() == "P" or h.strip() == "RT" or is_bool and any(
                    k in lh for k in ["gym", "squash", "diet", "wake", "sleep", "screentime", "quit scroll", "shower"]
                )) else "hours"
                columns.append({
                    "id": f"{user_id}-{n}-{h}".lower().replace(" ", "_"),
                    "name": h,
                    "type": col_type,
                    "weight": w,
                })

            # Read 7 day rows (rows 4..10)
            rows = []
            for day_i, day in enumerate(DAYS):
                excel_row = 4 + day_i
                values: dict[str, float] = {}
                for col_i, col in enumerate(columns):
                    excel_col = col_start + 1 + col_i
                    v = sheet_v.cell(row=excel_row, column=excel_col).value
                    if isinstance(v, (int, float)):
                        values[col["id"]] = float(v)
                    else:
                        values[col["id"]] = 0.0
                rows.append({"day": day, "values": values})

            tables.append({
                "userId": user_id,
                "userName": user_name,
                "columns": columns,
                "rows": rows,
            })

        weeks.append({
            "id": f"week-{n}",
            "weekNumber": n,
            "startDate": start_date.strftime("%Y-%m-%d"),
            "endDate": end_date.strftime("%Y-%m-%d"),
            "tables": tables,
        })

    seed = {
        "users": [
            {"id": "shazly", "name": "El Shazly", "color": "#7c3aed"},
            {"id": "sayed", "name": "El Sayed", "color": "#0ea5e9"},
        ],
        "weeks": weeks,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(seed, indent=2))
    print(f"Wrote {OUT_PATH} with {len(weeks)} weeks")


if __name__ == "__main__":
    main()
